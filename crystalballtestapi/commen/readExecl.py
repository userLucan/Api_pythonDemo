#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2020-07-18 16:50
# @Author  : 刘开


import openpyxl


class ReadExcel:
    """读取Excel"""
    def __init__(self, file_path, sheet_name=None):
        self.file_path = file_path
        self.sheet_name = sheet_name

    def open(self):
        """获取wb对象"""
        self.wb = openpyxl.load_workbook(self.file_path)

    def get_sheet(self):
        """
        获取sheet
        """
        self.open()
        sheet = self.wb[self.sheet_name]
        self.wb.close()
        return sheet

    def get_cell(self, row, col):
        """
        获取单元格对象
        """
        object = self.get_sheet().cell(row, col)
        self.wb.close()
        return object

    def get_cell_value(self, row, col):
        """
        获取某一个单元格的值
        """
        data = self.get_cell(row, col).value
        self.wb.close()
        return data

    def read(self):
        """读取所有数据"""
        # 运用列表推导式获取title
        headers = [i.value for i in self.get_sheet()[1]]
        row_data = []
        for row in range(2, self.get_sheet().max_row+1):
            data = []
            for col in range(1, self.get_sheet().max_column+1):
                data.append(self.get_cell_value(row, col))
            # 用zip函数将两个列表组合成字典
            data_info = dict(zip(headers, data))
            row_data.append(data_info)
        self.wb.close()
        return row_data

    def write(self, row, col, values):
        """
        回写数据
        :param row:
        :param col:
        :param values:
        :return:
        """
        self.get_cell(row, col).value = values
        self.save()
        return self.get_cell(row, col).value

    def save(self):
        """
        保存
        :return:
        """
        self.wb.save(self.file_path)